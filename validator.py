"""
검증 모듈: 계산값 vs 실제 시험값 비교
- 전제품_수정 24.07.02.xlsx의 시험 결과와 배합비 계산값 비교
"""
import re
import sys
import openpyxl
import pandas as pd
from pathlib import Path
from recipe_parser import parse_all_recipes
from calculator import calculate_nutrition
from ingredient_db import NUTRIENT_KEYS, NUTRIENT_LABELS

_lab_path = r"C:\Users\Moon\OneDrive\03. 양지희_품질\13. 영양성분\24년\전제품_수정 24.07.02.xlsx"
LAB_FILE = Path(_lab_path) if Path(_lab_path).exists() else None


def load_lab_results():
    """시험 결과 엑셀에서 제품별 영양성분 추출"""
    if LAB_FILE is None or not LAB_FILE.exists():
        return []
    wb = openpyxl.load_workbook(LAB_FILE, read_only=True, data_only=True)
    ws = wb['Sheet1']

    products = []
    for row_num in range(4, 93):  # 데이터 시작: 4행, ~92행
        name = ws.cell(row=row_num, column=2).value  # B: 제품명
        if not name:
            continue

        serving = ws.cell(row=row_num, column=10).value  # J: 내용량

        # 총 열량 파싱 (텍스트: "(520kcal)" 또는 "(5,060kcal)")
        kcal_raw = ws.cell(row=row_num, column=11).value  # K: 총 열량
        kcal = 0
        if kcal_raw:
            cleaned = str(kcal_raw).replace(',', '')
            match = re.search(r'[\d.]+', cleaned)
            if match:
                kcal = float(match.group())

        def safe_float(cell_val):
            if cell_val is None:
                return 0
            try:
                return float(cell_val)
            except (ValueError, TypeError):
                match = re.search(r'[\d.]+', str(cell_val))
                return float(match.group()) if match else 0

        # 1회 제공량 기준 열 (N열 = 14)
        per_serving_g = safe_float(ws.cell(row=row_num, column=14).value)  # N: ( )g당

        product = {
            "name": str(name).strip(),
            "serving_g": safe_float(serving),
            "per_serving_g": per_serving_g if per_serving_g > 0 else safe_float(serving),
            "lab": {
                "kcal": kcal,
                "sodium": safe_float(ws.cell(row=row_num, column=16).value),    # P
                "carb": safe_float(ws.cell(row=row_num, column=18).value),      # R
                "sugar": safe_float(ws.cell(row=row_num, column=20).value),     # T
                "fat": safe_float(ws.cell(row=row_num, column=24).value),       # X
                "trans_fat": safe_float(ws.cell(row=row_num, column=26).value), # Z
                "sat_fat": safe_float(ws.cell(row=row_num, column=27).value),   # AA
                "cholesterol": safe_float(ws.cell(row=row_num, column=29).value), # AC
                "protein": safe_float(ws.cell(row=row_num, column=31).value),   # AE
            }
        }
        products.append(product)

    wb.close()
    return products


def match_recipes_to_lab(recipes, lab_products):
    """레시피 이름과 시험 결과 이름 매칭"""
    matches = []

    for lab in lab_products:
        lab_name = lab["name"].strip()
        best_match = None
        best_score = 0

        for recipe in recipes:
            r_name = recipe["product_name"].strip()

            # 정확 매칭
            if lab_name == r_name:
                best_match = recipe
                best_score = 100
                break

            # 부분 매칭
            if lab_name in r_name or r_name in lab_name:
                score = len(set(lab_name) & set(r_name)) / max(len(set(lab_name)), len(set(r_name))) * 100
                if score > best_score:
                    best_match = recipe
                    best_score = score

        if best_match and best_score >= 50:
            matches.append({
                "lab": lab,
                "recipe": best_match,
                "match_score": best_score
            })

    return matches


def validate():
    """전체 검증 실행"""
    sys.stdout.reconfigure(encoding='utf-8')
    print("=" * 70)
    print("  영양성분 계산값 vs 시험값 비교 검증")
    print("=" * 70)

    # 1. 시험 결과 로드
    print("\n[1] 시험 결과 로딩...")
    lab_products = load_lab_results()
    print(f"    {len(lab_products)}개 제품 시험 결과 로드")

    # 2. 배합비 파싱
    print("\n[2] 배합비 파싱...")
    recipes = parse_all_recipes()
    print(f"    {len(recipes)}개 레시피 파싱 완료")

    # 3. 매칭
    print("\n[3] 제품 매칭...")
    matches = match_recipes_to_lab(recipes, lab_products)
    print(f"    {len(matches)}개 제품 매칭 성공")

    # 4. 계산 및 비교
    print("\n[4] 계산값 vs 시험값 비교")
    print("-" * 70)

    comparison_results = []

    for m in matches:
        lab = m["lab"]
        recipe = m["recipe"]
        serving_g = lab["serving_g"]  # 총 내용량(g)

        # 계산 (100g당 기준으로 통일)
        calc_result = calculate_nutrition(recipe["ingredients"])
        if calc_result is None:
            continue

        calc_per_100g = calc_result["per_100g"]

        # 시험값도 100g당으로 환산 (시험값은 총 내용량 기준)
        lab_vals_per_100g = {}
        if serving_g and serving_g > 0:
            for key in NUTRIENT_KEYS:
                lab_raw = lab["lab"].get(key, 0)
                lab_vals_per_100g[key] = lab_raw / serving_g * 100
        else:
            lab_vals_per_100g = lab["lab"]

        calc_vals = calc_per_100g
        lab_vals = lab_vals_per_100g

        errors = {}
        for key in NUTRIENT_KEYS:
            lab_val = lab_vals.get(key, 0) if key != "sodium" else lab_vals.get("sodium", 0)
            calc_val = calc_vals.get(key, 0)

            if lab_val > 0:
                error_pct = abs(calc_val - lab_val) / lab_val * 100
            elif calc_val > 0:
                error_pct = 100
            else:
                error_pct = 0

            errors[key] = {
                "lab": lab_val,
                "calc": calc_val,
                "error_pct": error_pct
            }

        comparison_results.append({
            "product": lab["name"],
            "coverage": calc_result["coverage"],
            "errors": errors,
            "unmapped_count": len(calc_result["unmapped"]),
        })

    # 5. 결과 출력
    if not comparison_results:
        print("  매칭된 제품이 없습니다.")
        return

    print(f"\n{'제품명':30s} {'매핑률':>6s} {'열량오차':>8s} {'탄수화물':>8s} {'단백질':>8s} {'지방':>8s} {'나트륨':>8s}")
    print("-" * 100)

    total_errors = {k: [] for k in ["kcal", "carb", "protein", "fat", "sodium"]}

    for r in comparison_results:
        name = r["product"][:28]
        cov = f"{r['coverage']:.0f}%"
        errs = r["errors"]

        kcal_e = f"{errs['kcal']['error_pct']:.0f}%"
        carb_e = f"{errs['carb']['error_pct']:.0f}%"
        prot_e = f"{errs['protein']['error_pct']:.0f}%"
        fat_e = f"{errs['fat']['error_pct']:.0f}%"
        sod_e = f"{errs['sodium']['error_pct']:.0f}%"

        print(f"  {name:28s} {cov:>6s} {kcal_e:>8s} {carb_e:>8s} {prot_e:>8s} {fat_e:>8s} {sod_e:>8s}")

        for key in total_errors:
            if errs[key]["lab"] > 0:
                total_errors[key].append(errs[key]["error_pct"])

    # 평균 오차
    print("-" * 100)
    avg_line = f"  {'평균 오차':28s} {'':>6s}"
    for key in ["kcal", "carb", "protein", "fat", "sodium"]:
        vals = total_errors[key]
        if vals:
            avg = sum(vals) / len(vals)
            avg_line += f" {avg:>7.1f}%"
        else:
            avg_line += f" {'N/A':>7s}%"
    print(avg_line)

    print(f"\n  총 비교 제품: {len(comparison_results)}개")
    print("=" * 70)

    return comparison_results


if __name__ == "__main__":
    validate()
