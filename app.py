"""
더브레드블루 영양성분 자동계산기 - Streamlit Web UI
"""
import sys
import os
import io
import re
sys.path.insert(0, os.path.dirname(__file__))

import streamlit as st
import pandas as pd
import openpyxl
from ingredient_db import INGREDIENT_DB, NUTRIENT_KEYS, NUTRIENT_LABELS, normalize_name, get_nutrient
from recipe_parser import parse_all_recipes, parse_recipe_file, parse_docx_file
from calculator import calculate_nutrition
from auto_lookup import lookup_ingredient, add_learned_ingredient, load_learned_db, find_similar_ingredients, get_api_key, save_api_key
from food_db_extended import EXTENDED_DB
from pathlib import Path

st.set_page_config(page_title="더브레드블루 영양성분 계산기", page_icon="🍞", layout="wide")

st.title("🍞 더브레드블루 영양성분 자동계산기")
st.caption("배합비 기반 9대 영양성분 자동 계산 시스템")

# ============================================================
# 공통 함수
# ============================================================
DAILY_REF = {
    "kcal": 2000, "carb": 324, "sugar": 100, "protein": 55,
    "fat": 54, "sat_fat": 15, "trans_fat": None,
    "cholesterol": 300, "sodium": 2000
}

def show_nutrition_result(result, serving_size, show_detail=True):
    """영양성분 계산 결과 표시 (공통)"""
    # 매핑률
    if result["coverage"] >= 90:
        st.success(f"DB 매핑률: {result['coverage']:.1f}% — 신뢰도 높음")
    elif result["coverage"] >= 70:
        st.warning(f"DB 매핑률: {result['coverage']:.1f}% — 일부 원재료 미매핑")
    else:
        st.error(f"DB 매핑률: {result['coverage']:.1f}% — 미매핑 원재료 多, 결과 부정확할 수 있음")

    # 영양성분 테이블: 제품중량당 + 100g당 + 1일 기준치
    rows = []
    for key in NUTRIENT_KEYS:
        label = NUTRIENT_LABELS[key]
        per100 = result["per_100g"][key]
        per_sv = result["per_serving"][key] if result["per_serving"] else per100

        # 1일 기준치 대비 (제품중량 기준)
        dv = DAILY_REF.get(key)
        if dv and isinstance(per_sv, (int, float)):
            pct = f"{per_sv / dv * 100:.0f}%"
        else:
            pct = "-"

        row = {"영양성분": label}
        if serving_size and serving_size != 100:
            row[f"제품중량({serving_size}g)당"] = round(per_sv, 1) if isinstance(per_sv, (int, float)) else per_sv
        row["100g당"] = round(per100, 1)
        row["1일 기준치 대비"] = pct
        rows.append(row)

    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    # 미매핑 원재료 — 자동 추천 + 학습
    if result["unmapped"]:
        st.divider()
        st.subheader(f"⚠️ 미매핑 원재료 ({len(result['unmapped'])}건)")
        st.caption("유사한 원재료를 선택하거나, 영양성분을 직접 입력하면 다음부터 자동 인식됩니다.")

        for um_name, um_qty in result["unmapped"]:
            um_pct = um_qty / result["total_weight"] * 100
            with st.expander(f"🔍 {um_name} — {um_qty:.1f}g ({um_pct:.1f}%)"):
                lookup = lookup_ingredient(um_name)
                suggestions = lookup.get("suggestions", [])

                if suggestions:
                    st.write("**유사 원재료 추천:**")
                    # 추천 목록 표시
                    options = ["직접 입력"] + [
                        f"{s['name']} (열량:{s['nutrition']['kcal']}kcal, 탄:{s['nutrition']['carb']}g, 단:{s['nutrition']['protein']}g, 지:{s['nutrition']['fat']}g)"
                        for s in suggestions[:5]
                    ]
                    choice = st.radio(
                        f"{um_name} 매핑 선택",
                        options,
                        key=f"radio_{um_name}",
                        label_visibility="collapsed"
                    )

                    if choice != "직접 입력":
                        idx = options.index(choice) - 1
                        selected_nutr = suggestions[idx]["nutrition"]
                        if st.button(f"✅ '{suggestions[idx]['name']}' 값으로 '{um_name}' 저장", key=f"save_{um_name}"):
                            add_learned_ingredient(um_name, selected_nutr, source="similar")
                            st.success(f"'{um_name}' → '{suggestions[idx]['name']}' 기준으로 저장 완료! 새로고침하면 반영됩니다.")
                            st.rerun()
                    else:
                        st.write("**직접 입력 (100g당):**")
                        col_a, col_b, col_c = st.columns(3)
                        with col_a:
                            m_kcal = st.number_input("열량(kcal)", min_value=0.0, value=0.0, key=f"kcal_{um_name}")
                            m_carb = st.number_input("탄수화물(g)", min_value=0.0, value=0.0, key=f"carb_{um_name}")
                            m_sugar = st.number_input("당류(g)", min_value=0.0, value=0.0, key=f"sugar_{um_name}")
                        with col_b:
                            m_protein = st.number_input("단백질(g)", min_value=0.0, value=0.0, key=f"protein_{um_name}")
                            m_fat = st.number_input("지방(g)", min_value=0.0, value=0.0, key=f"fat_{um_name}")
                            m_sat_fat = st.number_input("포화지방(g)", min_value=0.0, value=0.0, key=f"sat_{um_name}")
                        with col_c:
                            m_trans = st.number_input("트랜스지방(g)", min_value=0.0, value=0.0, key=f"trans_{um_name}")
                            m_chol = st.number_input("콜레스테롤(mg)", min_value=0.0, value=0.0, key=f"chol_{um_name}")
                            m_sodium = st.number_input("나트륨(mg)", min_value=0.0, value=0.0, key=f"sodium_{um_name}")

                        if st.button(f"💾 '{um_name}' 저장", key=f"manual_{um_name}"):
                            manual_nutr = {
                                "kcal": m_kcal, "carb": m_carb, "sugar": m_sugar,
                                "protein": m_protein, "fat": m_fat, "sat_fat": m_sat_fat,
                                "trans_fat": m_trans, "cholesterol": m_chol, "sodium": m_sodium
                            }
                            add_learned_ingredient(um_name, manual_nutr, source="manual")
                            st.success(f"'{um_name}' 저장 완료! 새로고침하면 반영됩니다.")
                            st.rerun()
                else:
                    st.write("유사 원재료를 찾지 못했습니다. 직접 입력해주세요.")
                    col_a, col_b, col_c = st.columns(3)
                    with col_a:
                        m_kcal = st.number_input("열량(kcal)", min_value=0.0, value=0.0, key=f"kcal_{um_name}")
                        m_carb = st.number_input("탄수화물(g)", min_value=0.0, value=0.0, key=f"carb_{um_name}")
                        m_sugar = st.number_input("당류(g)", min_value=0.0, value=0.0, key=f"sugar_{um_name}")
                    with col_b:
                        m_protein = st.number_input("단백질(g)", min_value=0.0, value=0.0, key=f"protein_{um_name}")
                        m_fat = st.number_input("지방(g)", min_value=0.0, value=0.0, key=f"fat_{um_name}")
                        m_sat_fat = st.number_input("포화지방(g)", min_value=0.0, value=0.0, key=f"sat_{um_name}")
                    with col_c:
                        m_trans = st.number_input("트랜스지방(g)", min_value=0.0, value=0.0, key=f"trans_{um_name}")
                        m_chol = st.number_input("콜레스테롤(mg)", min_value=0.0, value=0.0, key=f"chol_{um_name}")
                        m_sodium = st.number_input("나트륨(mg)", min_value=0.0, value=0.0, key=f"sodium_{um_name}")

                    if st.button(f"💾 '{um_name}' 저장", key=f"manual_{um_name}"):
                        manual_nutr = {
                            "kcal": m_kcal, "carb": m_carb, "sugar": m_sugar,
                            "protein": m_protein, "fat": m_fat, "sat_fat": m_sat_fat,
                            "trans_fat": m_trans, "cholesterol": m_chol, "sodium": m_sodium
                        }
                        add_learned_ingredient(um_name, manual_nutr, source="manual")
                        st.success(f"'{um_name}' 저장 완료! 새로고침하면 반영됩니다.")
                        st.rerun()


def parse_uploaded_file(uploaded_file):
    """업로드된 배합비 Excel에서 원재료+투입량 파싱"""
    wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
    all_results = []

    for ws in wb.worksheets:
        sheet_name = ws.title
        if '품목' in sheet_name:
            continue

        # 제품명 추출
        product_name = None
        for row_num in range(1, 4):
            cell_b = ws.cell(row=row_num, column=2).value
            if cell_b and isinstance(cell_b, str) and len(cell_b.strip()) > 1:
                val = cell_b.strip()
                if not any(k in val for k in ['단위', '규격', '원재료', '배합']):
                    product_name = val
                    break
        if not product_name:
            product_name = uploaded_file.name.replace('.xlsx', '')

        # 헤더 행 찾기
        header_row = None
        for row_num in range(1, 10):
            cell_a = ws.cell(row=row_num, column=1).value
            if cell_a and '원재료' in str(cell_a):
                header_row = row_num
                break

        if header_row is None:
            continue

        # 투입량 컬럼 찾기
        qty_col = 2
        for col in range(2, 8):
            cell = ws.cell(row=header_row, column=col).value
            if cell and ('1배합' in str(cell) or '배합' in str(cell)):
                qty_col = col
                break

        # 데이터 읽기
        ingredients = []
        for row_num in range(header_row + 1, 60):
            cell_a = ws.cell(row=row_num, column=1).value
            cell_qty = ws.cell(row=row_num, column=qty_col).value

            if cell_a is None:
                continue
            name = str(cell_a).strip()

            if any(k in name for k in ['합계', 'MEMO', 'memo', '소계', '총합', '개당', '판매']):
                break
            if len(name) < 2:
                continue
            if re.match(r'^[\d.,\s%()]+$', name):
                continue
            skip_kw = ['반죽', '성형', '굽기', '냉각', '분할', '℃', '오븐',
                       '뚜껑', '모양', '올려', '넣어', '섞어', '비고',
                       '단위', '규격', '입고', '원가', '백분율', '제품명']
            if any(k in name for k in skip_kw):
                continue
            if re.search(r'글루텐\s*\d', name) or name == '글루텐':
                continue

            qty = 0
            if cell_qty is not None:
                try:
                    qty = float(cell_qty)
                except (ValueError, TypeError):
                    match = re.search(r'[\d.]+', str(cell_qty))
                    if match:
                        qty = float(match.group())

            if qty > 0:
                ingredients.append((normalize_name(name), qty))

        if ingredients:
            all_results.append({
                "product_name": product_name,
                "sheet_name": sheet_name,
                "ingredients": ingredients,
                "total_weight": sum(q for _, q in ingredients)
            })

    wb.close()
    return all_results


# ============================================================
# 사이드바
# ============================================================
st.sidebar.header("메뉴")
mode = st.sidebar.radio("모드 선택", [
    "📤 배합비 파일 업로드 (신제품)",
    "📋 기존 배합비에서 계산",
    "✏️ 직접 입력 계산",
    "🔍 검증 (계산값 vs 시험값)",
    "📚 원재료 DB 조회",
])

# 사이드바 하단: DB 상태 + API 설정
st.sidebar.divider()
learned = load_learned_db()
st.sidebar.metric("기본 DB", f"{len(INGREDIENT_DB)}종")
st.sidebar.metric("확장 DB (식약처)", f"{len(EXTENDED_DB)}종")
st.sidebar.metric("학습된 원재료", f"{len(learned)}종")
st.sidebar.caption(f"총 {len(INGREDIENT_DB) + len(EXTENDED_DB) + len(learned)}종 검색 가능")

with st.sidebar.expander("⚙️ 식약처 API 설정"):
    current_key = get_api_key()
    if current_key:
        st.success("API키 등록됨")
        st.caption(f"키: {current_key[:8]}...")
    else:
        st.info("API키 없음 (유사명칭 검색만 사용)")
    new_key = st.text_input("API키 입력", type="password", placeholder="식품안전나라 인증키")
    if new_key and st.button("저장"):
        save_api_key(new_key)
        st.success("API키 저장 완료!")
        st.rerun()
    st.caption("[식품안전나라에서 API키 발급](https://www.foodsafetykorea.go.kr/apiMain.do)")

# ============================================================
# 1. 배합비 파일 업로드 (메인 기능)
# ============================================================
if mode == "📤 배합비 파일 업로드 (신제품)":
    st.header("배합비 파일 업로드")
    st.info("배합비 파일을 드래그앤드롭하세요. Excel(.xlsx) 또는 Word(.docx) 모두 지원합니다.")

    uploaded_files = st.file_uploader(
        "배합비 파일 (여러 개 가능)",
        type=["xlsx", "xls", "docx"],
        accept_multiple_files=True,
        help="Excel: A열 원재료명, B열 투입량(g) / Word: 원재료명+배합비율(%) 표"
    )

    if uploaded_files:
        for uploaded_file in uploaded_files:
            st.divider()
            st.subheader(f"📄 {uploaded_file.name}")

            try:
                if uploaded_file.name.endswith('.docx'):
                    # Word 파일: 배합비율(%) 표 파싱
                    docx_weight = None
                    # 파일명에서 중량 추출 시도
                    wt_m = re.search(r'(\d+)\s*[gG]', uploaded_file.name)
                    if wt_m:
                        docx_weight = float(wt_m.group(1))
                    else:
                        docx_weight = st.number_input(
                            "제품 총 중량 (g) — 배합비율(%)을 g으로 환산하는 데 필요",
                            min_value=1, value=100, step=10,
                            key=f"docx_wt_{uploaded_file.name}"
                        )
                    recipes = parse_docx_file(io.BytesIO(uploaded_file.read()), total_weight_g=docx_weight)
                else:
                    recipes = parse_uploaded_file(uploaded_file)
            except Exception as e:
                st.error(f"파일 파싱 실패: {e}")
                continue

            if not recipes:
                st.warning("원재료 데이터를 찾을 수 없습니다. 파일 양식을 확인해주세요.")
                st.caption("Excel: A열 = 원재료명, B열 = 투입량(g) / Word: 원재료명+배합비율(%) 표")
                continue

            # 레시피별 결과 표시 함수
            def show_recipe_result(recipe, file_name):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("제품명", recipe["product_name"])
                with col2:
                    st.metric("총 배합 중량", f"{recipe['total_weight']:.0f}g")
                with col3:
                    serving = st.number_input(
                        "제품 중량 (g)",
                        min_value=1,
                        value=100,
                        step=10,
                        key=f"serving_{file_name}_{recipe['sheet_name']}",
                        help="실제 판매 제품 1개 중량 (예: 40g, 85g, 120g, 200g 등)"
                    )

                ing_df = pd.DataFrame(recipe["ingredients"], columns=["원재료", "투입량(g)"])
                ing_df["비율(%)"] = (ing_df["투입량(g)"] / recipe["total_weight"] * 100).round(1)
                ing_df["DB매핑"] = ing_df["원재료"].apply(lambda x: "✅" if get_nutrient(x) else "❌")

                col_left, col_right = st.columns([1, 1])
                with col_left:
                    st.caption("배합비")
                    st.dataframe(ing_df, use_container_width=True, hide_index=True, height=400)
                with col_right:
                    st.caption("영양성분 계산 결과")
                    result = calculate_nutrition(recipe["ingredients"], serving_size_g=serving)
                    if result:
                        show_nutrition_result(result, serving)

            # 시트가 여러 개면 탭, 아니면 바로 표시
            if len(recipes) > 1:
                tabs = st.tabs([r["sheet_name"] for r in recipes])
                for tab, recipe in zip(tabs, recipes):
                    with tab:
                        show_recipe_result(recipe, uploaded_file.name)
            else:
                show_recipe_result(recipes[0], uploaded_file.name)

            # 여러 시트 비교 요약
            if len(recipes) > 1:
                st.divider()
                st.subheader("📊 시트별 비교 요약 (100g당)")
                summary_rows = []
                for recipe in recipes:
                    r = calculate_nutrition(recipe["ingredients"])
                    if r:
                        row = {"시트": recipe["sheet_name"], "제품명": recipe["product_name"],
                               "매핑률": f"{r['coverage']:.0f}%"}
                        for key in NUTRIENT_KEYS:
                            row[NUTRIENT_LABELS[key]] = round(r["per_100g"][key], 1)
                        summary_rows.append(row)
                if summary_rows:
                    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

    else:
        # 업로드 전 안내
        st.markdown("""
        ### 사용 방법
        1. 위 업로드 영역에 **배합비 파일**을 드래그앤드롭
        2. 자동으로 원재료와 투입량을 인식
        3. **9대 영양성분** 즉시 계산

        ### 지원하는 파일 형식
        - **Excel (.xlsx)**: A열 원재료명, B열 투입량(g), '원재료' 헤더 행 필요
        - **Word (.docx)**: 원재료명 + 배합비율(%) 표 (파일명에 중량 포함 시 자동 환산)
        - 기존 품질팀 배합비 양식 그대로 사용 가능
        - 여러 파일 동시 업로드 가능
        """)

        # 샘플 양식 다운로드
        sample_data = {
            "원재료": ["통밀가루", "정제수", "현미유", "하얀설탕", "정제소금", "드라이이스트"],
            "1배합": [500, 350, 30, 20, 8, 5]
        }
        sample_df = pd.DataFrame(sample_data)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            sample_df.to_excel(writer, index=False, sheet_name='Sheet1')
        st.download_button(
            "📥 샘플 양식 다운로드",
            data=buffer.getvalue(),
            file_name="배합비_샘플양식.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document"
        )

# ============================================================
# 2. 기존 배합비에서 계산
# ============================================================
elif mode == "📋 기존 배합비에서 계산":
    st.header("기존 배합비에서 영양성분 계산")

    @st.cache_data
    def load_recipes():
        return parse_all_recipes()

    with st.spinner("배합비 파일 파싱 중..."):
        recipes = load_recipes()

    if not recipes:
        st.info("로컬 배합비 폴더에 접근할 수 없습니다. '배합비 파일 업로드' 메뉴를 사용해주세요.")
        st.stop()

    st.success(f"{len(recipes)}개 레시피 로드 완료")

    # 검색 기능
    search_term = st.text_input("제품명 검색", placeholder="예: 통밀, 스콘, 크림빵...")
    product_names = [r["product_name"] for r in recipes]
    if search_term:
        product_names = [n for n in product_names if search_term in n]

    if not product_names:
        st.warning("검색 결과가 없습니다.")
    else:
        selected = st.selectbox("제품 선택", product_names, index=0)
        recipe = next(r for r in recipes if r["product_name"] == selected)

        col1, col2 = st.columns(2)
        with col1:
            serving_size = st.number_input("제품 중량 (g)", min_value=1, value=100, step=10, help="실제 판매 제품 1개 중량")
        with col2:
            st.metric("총 배합 중량", f"{recipe['total_weight']:.0f}g")

        # 배합비
        ing_df = pd.DataFrame(recipe["ingredients"], columns=["원재료", "투입량(g)"])
        ing_df["비율(%)"] = (ing_df["투입량(g)"] / recipe["total_weight"] * 100).round(1)
        ing_df["DB매핑"] = ing_df["원재료"].apply(lambda x: "✅" if get_nutrient(x) else "❌")

        col_left, col_right = st.columns([1, 1])
        with col_left:
            st.subheader("배합비")
            st.dataframe(ing_df, use_container_width=True, hide_index=True)
        with col_right:
            st.subheader("영양성분 계산 결과")
            result = calculate_nutrition(recipe["ingredients"], serving_size_g=serving_size)
            if result:
                show_nutrition_result(result, serving_size)

# ============================================================
# 3. 직접 입력 계산
# ============================================================
elif mode == "✏️ 직접 입력 계산":
    st.header("원재료 직접 입력")
    st.info("원재료명과 투입량(g)을 입력하세요. 행 추가는 하단 + 버튼")

    if "manual_ingredients" not in st.session_state:
        st.session_state.manual_ingredients = pd.DataFrame({
            "원재료명": ["통밀가루", "정제수", "현미유", "정제소금", "드라이이스트"],
            "투입량(g)": [500.0, 350.0, 30.0, 8.0, 5.0]
        })

    edited_df = st.data_editor(
        st.session_state.manual_ingredients,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "원재료명": st.column_config.SelectboxColumn(
                "원재료명",
                options=sorted(INGREDIENT_DB.keys()),
                required=True,
            ),
            "투입량(g)": st.column_config.NumberColumn("투입량(g)", min_value=0, step=1.0),
        }
    )

    serving_size = st.number_input("제품 중량 (g)", min_value=1, value=100, step=10, help="실제 판매 제품 1개 중량")

    if st.button("계산하기", type="primary"):
        ingredients = []
        for _, row in edited_df.iterrows():
            name = row["원재료명"]
            qty = row["투입량(g)"]
            if name and qty and qty > 0:
                ingredients.append((normalize_name(str(name)), float(qty)))

        if ingredients:
            result = calculate_nutrition(ingredients, serving_size_g=serving_size)
            if result:
                st.subheader("계산 결과")
                show_nutrition_result(result, serving_size)
        else:
            st.error("원재료를 입력해주세요.")

# ============================================================
# 4. 검증
# ============================================================
elif mode == "🔍 검증 (계산값 vs 시험값)":
    st.header("계산값 vs 시험값 비교 검증")
    st.caption("기존 시험 의뢰 결과와 배합비 계산값의 오차를 비교합니다.")

    if st.button("검증 실행", type="primary"):
        with st.spinner("배합비 파싱 + 시험결과 로딩 중..."):
            from validator import load_lab_results, match_recipes_to_lab

            lab_products = load_lab_results()
            recipes = parse_all_recipes()
            matches = match_recipes_to_lab(recipes, lab_products)

        st.info(f"시험결과 {len(lab_products)}개 / 배합비 {len(recipes)}개 / 매칭 {len(matches)}개")

        if matches:
            rows = []
            for m in matches:
                lab = m["lab"]
                recipe = m["recipe"]
                serving_g = lab["serving_g"]
                calc_result = calculate_nutrition(recipe["ingredients"])

                if calc_result is None:
                    continue

                calc_100g = calc_result["per_100g"]
                lab_100g = {}
                if serving_g and serving_g > 0:
                    for key in NUTRIENT_KEYS:
                        lab_100g[key] = lab["lab"].get(key, 0) / serving_g * 100
                else:
                    lab_100g = lab["lab"]

                row = {"제품명": lab["name"], "매핑률": f"{calc_result['coverage']:.0f}%"}

                for key, label in [("kcal", "열량"), ("carb", "탄수화물"), ("protein", "단백질"), ("fat", "지방"), ("sodium", "나트륨")]:
                    lv = lab_100g.get(key, 0)
                    cv = calc_100g.get(key, 0)
                    err = abs(cv - lv) / lv * 100 if lv > 0 else 0
                    row[f"{label} 시험"] = round(lv, 1)
                    row[f"{label} 계산"] = round(cv, 1)
                    row[f"{label} 오차%"] = round(err, 1)

                rows.append(row)

            if rows:
                df = pd.DataFrame(rows)
                st.dataframe(df, use_container_width=True, hide_index=True)

                st.subheader("평균 오차율 (100g당 기준)")
                cols = st.columns(5)
                for i, key_label in enumerate(["열량", "탄수화물", "단백질", "지방", "나트륨"]):
                    col_name = f"{key_label} 오차%"
                    vals = [r[col_name] for r in rows if isinstance(r[col_name], (int, float)) and r[col_name] < 200]
                    if vals:
                        avg = sum(vals) / len(vals)
                        cols[i].metric(key_label, f"{avg:.1f}%")
        else:
            st.warning("매칭된 제품이 없습니다.")

# ============================================================
# 5. 원재료 DB 조회
# ============================================================
elif mode == "📚 원재료 DB 조회":
    st.header("원재료 영양성분 DB")

    search = st.text_input("원재료 검색", placeholder="예: 통밀가루, 버터, 설탕...")

    # 전체 DB 통합 (기본 + 확장 + 학습)
    all_db = {}
    for name, nutrients in INGREDIENT_DB.items():
        all_db[name] = ("기본", nutrients)
    for name, nutrients in EXTENDED_DB.items():
        if name not in all_db:
            all_db[name] = ("확장", nutrients)
    for name, data in load_learned_db().items():
        if name not in all_db:
            all_db[name] = ("학습", data["nutrition"])

    rows = []
    for name, (source, nutrients) in all_db.items():
        if search and search not in name:
            continue
        row = {"원재료명": name, "출처": source}
        for key in NUTRIENT_KEYS:
            row[NUTRIENT_LABELS[key]] = nutrients.get(key, 0)
        rows.append(row)

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)
    st.caption(f"총 {len(df)}개 원재료 (기본 {len(INGREDIENT_DB)} + 확장 {len(EXTENDED_DB)} + 학습 {len(load_learned_db())})")

    with st.expander("📝 새 원재료 추가 요청"):
        st.write("DB에 없는 원재료는 아래에 입력해주세요.")
        new_name = st.text_input("원재료명")
        if new_name:
            if new_name in INGREDIENT_DB or normalize_name(new_name) in INGREDIENT_DB:
                st.success(f"'{new_name}'은(는) 이미 DB에 등록되어 있습니다.")
            else:
                st.info(f"'{new_name}'은(는) 아직 DB에 없습니다. ingredient_db.py에 추가해주세요.")
